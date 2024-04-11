from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

# 統合されたExcelファイルのパス
#ダウンロードしてから使用する場合、ここを編集してください！！
integrated_file_path = ''

# 統合されたExcelファイルを開く
wb = load_workbook(integrated_file_path)

# 融合されたグラフを作成
combined_chart = ScatterChart()
combined_chart.title = "Combined Scatter Chart"
combined_chart.x_axis.title = 'E column values'
combined_chart.y_axis.title = 'D column values'
combined_chart.legend.position = 'r'  # 凡例を右側に配置

# 統合されたExcelファイルの各シートを処理してグラフを作成
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # グラフのデータ範囲を指定（29行目から最終行までのD列とE列）
    min_row = 29
    max_row = ws.max_row
    x_values = Reference(ws, min_col=5, min_row=min_row, max_row=max_row)  # E列
    y_values = Reference(ws, min_col=4, min_row=min_row, max_row=max_row)  # D列

    # シートごとの散布図のデータ系列を作成して融合グラフに追加
    series = Series(y_values, x_values, title=sheet_name)
    series.marker.symbol = "circle"  # 点を丸で表示
    series.graphicalProperties.line.noFill = True  # 線を表示しない
    combined_chart.series.append(series)

# 最初のシートに融合されたグラフを追加
first_sheet = wb.worksheets[0]
first_sheet.add_chart(combined_chart, "G10")  # グラフを配置するセルを指定（例: G10）

# 変更を保存
wb.save(integrated_file_path)

print(f"{integrated_file_path} に融合された散布図が追加されました。")