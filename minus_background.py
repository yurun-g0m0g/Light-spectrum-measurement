import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 統合されたExcelファイルのパス
integrated_file_path = '/Users/sidareyanagi542/Desktop/授業資料/4年/研究室/実験/0405/統合ファイル.xlsx'

# 参照するExcelファイルのパス
reference_file_path = '/Users/sidareyanagi542/Desktop/授業資料/4年/研究室/実験/0405/background.xlsx'

# 統合されたExcelファイルを開く
wb = load_workbook(integrated_file_path)

# 参照するExcelファイルを開く
wb_ref = load_workbook(reference_file_path)
ws_ref = wb_ref.active  # 参照ファイルのアクティブなシートを使用

# 統合されたExcelファイルの各シートを処理
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # D列とE列の29行目以下の各セルについて処理
    for row in range(29, ws.max_row + 1):  # 29行目から最終行まで
        # D列: 参照ファイルの値を引く処理
        d_cell = ws[f"D{row}"]
        d_ref_cell = ws_ref[f"D{row}"]
        try:
            d_cell_value = float(d_cell.value) if d_cell.value is not None and d_cell.value != '' else 0
            d_ref_cell_value = float(d_ref_cell.value) if d_ref_cell.value is not None and d_ref_cell.value != '' else 0
            d_cell.value = d_cell_value - d_ref_cell_value
        except ValueError:
            print(f"変換エラー: シート '{sheet_name}', セル D{row} は数値に変換できません。")

        # E列: 1239.8をC列の値で割った結果を設定
        c_cell = ws[f"C{row}"]
        e_cell = ws[f"E{row}"]
        try:
            c_cell_value = float(c_cell.value) if c_cell.value is not None and c_cell.value != '' else 1  # 0による除算を避ける
            if c_cell_value != 0:  # 0除算の防止
                e_cell.value = 1239.8 / c_cell_value
            else:
                e_cell.value = "Error: Division by zero"
        except ValueError:
            e_cell.value = "Error: Invalid C column value"

# 変更を保存
wb.save(integrated_file_path)

print(f"{integrated_file_path} の各シートのD列とE列の29行目以下が更新されました。")